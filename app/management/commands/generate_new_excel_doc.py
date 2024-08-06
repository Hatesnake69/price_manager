import io

import openpyxl
from django.core.management.base import BaseCommand
from openpyxl.styles import PatternFill

from app.models import KaspiGoodsModel
from app.utils.send_docs_via_tg import send_docs


class Command(BaseCommand):
    help = 'Создаёт документ с обновленными ценами'

    def handle(self, *args, **kwargs):

        prices_changed_data_fill = PatternFill(
            start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
        )    # Зеленая заливка для данных
        min_price_too_high_fill = PatternFill(
            start_color="FF0000", end_color="FF0000", fill_type="solid"
        )  # Красная заливка для данных


        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'KaspiGoods'

        # Запись заголовков столбцов
        sheet.append([
            "sku",
            "model",
            "brand",
            "price",
            "pp1",
            "pp2",
            "pp3",
            "pp4",
            "pp5",
            "preorder",
        ])
        for row_idx, good in enumerate(KaspiGoodsModel.objects.order_by('id'), start=2):
            sheet.append([
                good.sku,
                good.model,
                good.brand,
                good.price,
                good.pp1,
                good.pp2,
                good.pp3,
                good.pp4,
                good.pp5,
                good.preorder,
            ])
            if all((good.prev_price, good.price != good.prev_price)):
                for cell in sheet[row_idx]:
                    cell.fill = prices_changed_data_fill
            elif good.min_price_too_high_flag:
                for cell in sheet[row_idx]:
                    cell.fill = min_price_too_high_fill

        workbook.save(output)
        output.seek(0)
        send_docs(docs=[output])
        self.stdout.write(self.style.SUCCESS('Команда выполнена успешно!'))

